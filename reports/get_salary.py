from bd.model import Clients, Document
from arrow import utcnow, get
from bd.model import Session, Clients, Documents
from .util import xls_to_json_format_change, json_to_xls_format_change_
from pprint import pprint
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, get_column_letter
import sys  # Импортируем модуль sys для получения информации о текущем исключении
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)

name = "🤑🤑🤑 Зарплата ➡️".upper()
desc = "Загружает данне из xls в базу"
mime = "file_5"


class FileCRMInput:
    name = "Файл CRM"
    desc = "Отправте файл CRM  в формате xlsx"
    type = "FILE"


class FileCdekInput:
    name = "Файл cdek"
    desc = "Отправте файл cdek в формате xlsx"
    type = "FILE"


class FileProvidedFileInput:
    name = "Файл provided file"
    desc = "Отправте предоставляемый файл в формате xlsx"
    type = "FILE"


def get_inputs(session: Session):
    return {
        "CRM": FileCRMInput,
        "Cdek": FileCdekInput,
        "ProvidedFile": FileProvidedFileInput,
    }


def generate(session: Session):
    try:
        logging.info("Начало генерации данных")

        # Получаем параметры сессии
        params = session.params["inputs"]["0"]

        logging.debug(f"Параметры сессии: {params}")
        # Создаем словарь для данных CRM
        crm_dict = {}

        # Создаем словарь CRM, где ключ - оператор, значение - список ID заказов
        for item in params["CRM"]:
            operator = item.get("Оператор")
            if operator not in crm_dict:
                crm_dict[operator] = []
            crm_dict[operator].append(item.get("ID"))

        logging.debug(f"CRM данные: {crm_dict}")

        # Получаем ключи первого элемента списка Cdek
        keys = list(params["Cdek"][0].keys())

        # Получаем первый и второй ключи из Cdek
        key1 = keys[0]
        key2 = keys[1]

        # Создаем словарь  для каждого заказа из Cdek
        cdek_date = {
            item[key1]: item[key2]
            for item in params["Cdek"]
            if None not in (item.get(key1), item.get(key2))
        }

        logging.debug(f"Cdek данные: {cdek_date}")

        # Создаем словари для сумм продаж, списка заказов и данных о заказах
        dict_sales = {}
        dict_order = {}
        dict_order_list_ = {}

        # Обрабатываем данные для каждого пользователя CRM
        for user, order_list in crm_dict.items():
            sum_ = 0
            order_ = {}
            order_list_ = []

            # Обходим список заказов для каждого оператора
            for order in order_list:
                # Проверяем наличие данных о заказе в Cdek
                if order in cdek_date:
                    order_list_.append(order)
                    order_.update({str(order): cdek_date.get(order, 0)})

                sum_ += cdek_date.get(order, 0)

            # Заполняем словари данными
            dict_sales.update({user: sum_})
            dict_order.update({user: order_})
            dict_order_list_.update({user: order_list_})

        logging.debug(f"Словарь продаж: {dict_sales}")
        logging.debug(f"Словарь заказов: {dict_order}")

        # Извлекаем данные о файлах из параметров сессии ProvidedFile
        provided_file_ = params["ProvidedFile"]

        provided_file = []

        oldest_salesman_data = {}

        preoldest_salesman_list = ["POD", "SKP", "UDL", "M31"]

        # pprint(provided_file_)

        # Заменяем значения None на 0 в каждом словаре
        # Итерируемся по элементам списка provided_file_
        for item in provided_file_:
            # Проверяем, есть ли текущий сотрудник в списке preoldest_salesman_list
            if item["Сотрудник"] in preoldest_salesman_list:
                # Если сотрудник является одним из старейших, создаем новый словарь new_item,
                # где заменяем значения None на 0
                new_item = {
                    key: (value if value is not None else 0)
                    for key, value in item.items()
                }

                # Обновляем словарь oldest_salesman_data, используя имя сотрудника в качестве ключа
                # и новый словарь new_item в качестве значения
                oldest_salesman_data.update({item["Сотрудник"]: new_item})
            else:
                # Если сотрудник не является старейшим, также создаем новый словарь new_item,
                # где заменяем значения None на 0
                new_item = {
                    key: (value if value is not None else 0)
                    for key, value in item.items()
                }
                # Добавляем новый словарь new_item в исходный список provided_file_
                provided_file.append(new_item)

        logging.debug(f"Файл данных: {provided_file}")
        logging.debug(f"Данные старейших сотрудников: {oldest_salesman_data}")

        # Инициализируем переменные для общей суммы и списка данных
        total_date = []
        message_date = []
        total_sum = 0

        # Обходим данные из файла
        for item in provided_file:
            if item["Сотрудник"] != 0:
                try:
                    # Вычисляем общую сумму для каждого сотрудника
                    total = (
                        (
                            Decimal(
                                str(dict_sales.get(item["Сотрудник"], 0))
                            )  # Получаем продажи для текущего сотрудника из словаря dict_sales
                            * Decimal(str(item["%"]))  # Процент от продаж
                        )
                        + Decimal(str(item["Оклад"]))  # Оклад сотрудника
                        + Decimal(str(item["Отпускные"]))  # Сумма отпускных
                        - Decimal(str(item["Офчасть"]))  # Сумма официальной части
                        - Decimal(str(item["Долг"]))  # Сумма долгов
                        + Decimal(str(item["доп премия"]))  # Дополнительная премия
                    )
                    total = total.quantize(Decimal("0.00"))  # Дополнительная премия

                    total_sum += total  # Добавляем текущую сумму к общей сумме

                    # Формируем результаты для записи
                    result = {
                        "Сотрудник": item["Сотрудник"],  # Идентификатор сотрудника
                        "Сумма": Decimal(dict_sales.get(item["Сотрудник"], 0)).quantize(
                            Decimal("0.00")
                        ),  # Сумма продаж для сотрудника из словаря dict_sales
                        "%": (Decimal(item["%"]) * 100).quantize(
                            Decimal("0.00")
                        ),  # Процент от продаж
                        "Итог%": (
                            Decimal(dict_sales.get(item["Сотрудник"], 0))
                            * Decimal(
                                item["%"]
                            )  # Итоговая сумма по процентам от продаж
                        ).quantize(Decimal("0.00")),
                        "Оклад": Decimal(item["Оклад"]).quantize(
                            Decimal("0.00")
                        ),  # Оклад сотрудника
                        "Отпускные": Decimal(item["Отпускные"]).quantize(
                            Decimal("0.00")
                        ),  # Сумма отпускных
                        "Офчасть": Decimal(item["Офчасть"]).quantize(
                            Decimal("0.00")
                        ),  # Сумма официальной части
                        "Долг": Decimal(item["Долг"]).quantize(
                            Decimal("0.00")
                        ),  # Сумма долга
                        "доп премия": Decimal(item["доп премия"]).quantize(
                            Decimal("0.00")  # Дополнительная премия
                        ),
                        "Итог": total,  # Общая сумма для сотрудника
                    }

                    result_message = {
                        "Сотрудник:": item["Сотрудник"],  # Идентификатор сотрудника
                        "Сумма:": "{}₱".format(
                            result["Сумма"]
                        ),  # Сумма продаж для сотрудника из словаря dict_sales
                        "Процент:": "{}%".format(result["%"]),  # Процент от продаж
                        "Итог %:": "{}₱".format(
                            result["Итог%"]
                        ),  # Итоговая сумма по процентам от продаж
                        "Оклад:": "{}₱".format(result["Оклад"]),  # Оклад сотрудника
                        "Отпускные:": "{}₱".format(
                            result["Отпускные"]
                        ),  # Сумма отпускных
                        "Офчасть:": "{}₱".format(
                            result["Офчасть"]
                        ),  # Сумма официальной части
                        "Долг:": "{}₱".format(result["Долг"]),  # Сумма долга
                        "Доп премия:": "{}₱".format(
                            result["доп премия"]
                        ),  # Дополнительная премия
                        "Итог:": "{}₱".format(total),  # Общая сумма для сотрудника
                    }

                    total_date.append(
                        result
                    )  # Добавляем результаты в список total_date

                    message_date.append(
                        result_message
                    )  # Добавляем результаты в список message_date

                    result_up = {}  # Создаем пустой словарь result_up
                    result_up.update(
                        result
                    )  # Обновляем словарь result_up данными из result

                    # Форматируем заказы для записи
                    format_order = {}
                    format_order_messange = {}
                    if item["Сотрудник"] in dict_order:
                        for k, v in dict_order[item["Сотрудник"]].items():
                            sum_order = Decimal(float(v)).quantize(Decimal("0.00"))
                            # pprint(float(sum_order))
                            format_order.update({str(k): sum_order})
                            format_order_messange.update(
                                {"№{}:".format(k): "{}₱".format(sum_order)}
                            )

                    result.update(format_order)
                    result_message.update(format_order_messange)

                    # Обновляем дополнительные данные в результате
                    result_up.update(
                        {
                            "closeDate": utcnow().shift(hours=3).isoformat(),
                            "order_list": dict_order_list_.get(item["Сотрудник"], []),
                            "order": dict_order.get(item["Сотрудник"], {}),
                        }
                    )

                    # Конвертация Decimal в float
                    result_up_float = {
                        k: float(v) if isinstance(v, Decimal) else v
                        for k, v in result_up.items()
                    }
                    # pprint(result_up_float)

                    # Обновление документов в MongoDB
                    Documents.objects(closeDate=result_up["closeDate"]).update(
                        **result_up_float, upsert=True
                    )
                except Exception as e:
                    logging.error(
                        f"Ошибка: {e} на строке {sys.exc_info()[-1].tb_lineno}"
                    )

        prefix = ["POD", "SKP", "UDL", "M31"]

        # Создаем словарь, где ключи - префиксы, значения - список элементов
        prefix_dict: dict = {p: [] for p in prefix}
        # pprint(total_date)
        # Заполняем префиксный словарь
        for item in total_date:
            for p in prefix:
                if item["Сотрудник"].startswith(
                    p
                ):  # Проверяем, начинается ли значение ключа "Сотрудник" с текущего префикса
                    prefix_dict[p].append(
                        item
                    )  # Если да, добавляем элемент в список, соответствующий текущему префиксу

        for (
            pref,
            list_,
        ) in prefix_dict.items():  # Перебираем ключи и значения в словаре prefix_dict
            sum_items = 0
            for item in list_:
                sum_items += Decimal(item["Сумма"]).quantize(
                    Decimal("0.00")
                )  # Считаем сумму значений ключа "Сумма" для всех элементов списка

            # Конвертировать операнды в Decimal перед выполнением операции
            # pprint(type(sum_items))
            # pprint(sum_items * Decimal(str(oldest_salesman_data[pref]["%"])))

            sum_items_decimal = Decimal(sum_items)
            # pprint(oldest_salesman_data.get(pref, {}).get("%", 0))
            total_ = Decimal(
                (
                    sum_items
                    * Decimal(str(oldest_salesman_data.get(pref, {}).get("%", 0)))
                )
                + oldest_salesman_data.get(pref, {}).get("Оклад", 0)
                + oldest_salesman_data.get(pref, {}).get("Отпускные", 0)
                - oldest_salesman_data.get(pref, {}).get("Офчасть", 0)
                - oldest_salesman_data.get(pref, {}).get("Долг", 0)
                + oldest_salesman_data.get(pref, {}).get("доп премия", 0)
            ).quantize(Decimal("0.00"))

            resut = {
                "%": Decimal(
                    oldest_salesman_data.get(pref, {}).get("%", 0) * 100
                ).quantize(Decimal("0.00")),
                "Долг": Decimal(
                    oldest_salesman_data.get(pref, {}).get("Долг", 0)
                ).quantize(Decimal("0.00")),
                "Итог": total_,
                "Итог%": Decimal(
                    sum_items * oldest_salesman_data.get(pref, {}).get("Итог%", 0)
                ).quantize(Decimal("0.00")),
                "Оклад": Decimal(
                    oldest_salesman_data.get(pref, {}).get("Оклад", 0)
                ).quantize(Decimal("0.00")),
                "Отпускные": Decimal(
                    oldest_salesman_data.get(pref, {}).get("Отпускные", 0)
                ).quantize(Decimal("0.00")),
                "Офчасть": Decimal(
                    oldest_salesman_data.get(pref, {}).get("Офчасть", 0)
                ).quantize(Decimal("0.00")),
                "Сотрудник": pref,
                "Сумма": Decimal(sum_items).quantize(Decimal("0.00")),
                "доп премия": Decimal(
                    oldest_salesman_data.get(pref, {}).get("доп премия", 0)
                ).quantize(Decimal("0.00")),
            }
            list_.append(resut)  # Изменил способ создания словаря
            total_date.append(resut)

            result_message = {
                "Сотрудник:": pref,  # Идентификатор сотрудника
                "Сумма:": "{}₱".format(
                    Decimal(sum_items).quantize(Decimal("0.00"))
                ),  # Сумма продаж для сотрудника из словаря dict_sales
                "Процент:": "{}%".format(
                    Decimal(
                        oldest_salesman_data.get(pref, {}).get("%", 0) * 100
                    ).quantize(Decimal("0.00"))
                ),  # Процент от продаж
                "Итог %:": "{}₱".format(
                    Decimal(
                        sum_items * oldest_salesman_data.get(pref, {}).get("Итог%", 0)
                    ).quantize(Decimal("0.00"))
                ),
                "Оклад:": "{}₱".format(
                    Decimal(
                        oldest_salesman_data.get(pref, {}).get("Оклад", 0)
                    ).quantize(Decimal("0.00"))
                ),  # Оклад сотрудника
                "Отпускные:": "{}₱".format(
                    Decimal(
                        oldest_salesman_data.get(pref, {}).get("Отпускные", 0)
                    ).quantize(Decimal("0.00"))
                ),  # Сумма отпускных
                "Офчасть:": "{}₱".format(
                    Decimal(
                        oldest_salesman_data.get(pref, {}).get("Офчасть", 0)
                    ).quantize(Decimal("0.00"))
                ),  # Сумма официальной части
                "Долг:": "{}₱".format(
                    Decimal(
                        oldest_salesman_data.get(pref, {}).get("доп премия", 0)
                    ).quantize(Decimal("0.00"))
                ),  # Сумма долга
                "Доп премия:": "{}₱".format(
                    Decimal(
                        oldest_salesman_data.get(pref, {}).get("доп премия", 0)
                    ).quantize(Decimal("0.00"))
                ),
                "Итог:": "{}₱".format(total_),  # Общая сумма для сотрудника
            }
            message_date.append(result_message)
        # pprint(prefix_dict)

        # Форматируем данные для экспорта в Excel и добавляем в список книг
        books = []
        for k, v in prefix_dict.items():
            if len(v) > 0:
                books.append(json_to_xls_format_change_(v))
        message_date.append(
            {
                "closeDate": utcnow().shift(hours=3).isoformat()[:10],
                "Итог": total_sum,
            }
        )
        logging.info("Генерация данных завершена")

        # Возвращаем данные и книги Excel
        return message_date, books
    except Exception as e:
        logging.error(f"Ошибка: {e} на строке {sys.exc_info()[-1].tb_lineno}")
