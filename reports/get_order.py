from bd.model import Clients, Documents
from arrow import utcnow, get
from bd.model import Session, Clients
from .util import xls_to_json_format_change
from pprint import pprint
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, get_column_letter


name = "📦 Get Order ➡️"
desc = "Загружает данне из xls в базу"
mime = "text"


class OrderInput:
    name = "Файл"
    desc = "Напишите номер заказа"
    type = "MESSAGE"


def get_inputs(session: Session):
    return {"order": OrderInput}


def generate(session: Session):
    params = session.params["inputs"]["0"]

    order_ = int(params["order"])
    pprint(type(order_))
    try:

        document = Documents.objects(order_list__in=[order_])
        report_date = []
        for doc in document:
            dic_doc: dict = doc["order"]
            dic_doc.update(
                {
                    "closeDate": doc["closeDate"][:10],
                    "Сотрудник": doc["Сотрудник"],
                    "Сумма": doc["Сумма"],
                    "%": doc["%"],
                    "Итог%": doc["Итог%"],
                    "Оклад": doc["Оклад"],
                    "Отпускные": doc["Отпускные"],
                    "Офчасть": doc["Офчасть"],
                    "Долг": doc["Долг"],
                    "доп премия": doc["доп премия"],
                    "Итог": doc["Итог"],
                }
            )
            report_date.append(dic_doc)
        pprint(report_date)
    except Exception as e:
        print(f"Error sending messages: {e}")
    return report_date
