import PyPDF2
import re
import json
import requests
from arrow import utcnow, get
from typing import List, Tuple
from pprint import pprint
from openpyxl.utils import get_column_letter
import io
from pprint import pprint
import zipfile
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
import re
import json
import requests
import PyPDF2
import traceback
import json
import re
import time
import rarfile
import io
import pdfplumber


# import psutil  # Для мониторинга загрузки процессора
from pprint import pprint

import logging

logger = logging.getLogger(__name__)


def format_message_list2(obj):
    text = ""  # Создаем пустую строку, в которую будем добавлять текст
    messages = []  # Создаем пустой список для хранения сообщений

    if len(obj) > 0:  # Проверяем, что входной объект не пуст
        logger.debug("Processing non-empty object")
        for k, v in obj.items():  # Проходим по ключам и значениям в объекте
            key = str(k)  # Преобразуем ключ в строку
            val = str(v)  # Преобразуем значение в строку
            total_len = len(key) + len(val)  # Вычисляем общую длину ключа и значения
            pad = 31 - total_len % 31  # Вычисляем количество пробелов для выравнивания

            text += key  # Добавляем ключ в текст

            if pad > 0:
                text += " " * pad  # Добавляем пробелы для выравнивания

            if total_len > 31:
                text += " " * 2  # Добавляем двойные пробелы, если общая длина больше 31

            text += str(v)  # Добавляем значение в текст
            text += "\n"  # Добавляем символ новой строки

        # Разбиваем текст на части, если он слишком большой
        index = 0
        size = 4000
        while len(text) > 0:
            part = text[index : index + size]  # Вырезаем часть текста заданного размера
            index = part.rfind("\n")  # Находим последний символ новой строки в части
            if index == -1:
                index = len(
                    text
                )  # Если символ новой строки не найден, используем конец текста
            part = text[0:index]  # Выбираем часть текста до символа новой строки
            messages.append(
                "```\n" + part + "\n```"
            )  # Добавляем часть текста в список сообщений
            text = text[
                index:
            ].strip()  # Удаляем обработанную часть из текста и убираем пробелы

        logger.debug("Finished formatting messages")
    else:
        logger.debug("Input object is empty")

    return messages  # Возвращаем список сообщений


def format_message_list4(obj):
    text = ""  # Создаем пустую строку для хранения текста сообщений.
    messages = []  # Создаем пустой список для хранения отформатированных сообщений.

    if len(obj) > 0:  # Проверяем, есть ли объекты в списке.
        logger.debug("Processing non-empty object")
        for i in obj:  # Проходим по каждому объекту в списке.
            for k, v in i.items():  # Проходим по каждой паре ключ-значение в объекте.
                key = str(k)  # Преобразуем ключ в строку.
                val = str(v)  # Преобразуем значение в строку.
                total_len = len(key) + len(
                    val
                )  # Вычисляем общую длину ключа и значения.
                pad = (
                    30 - total_len % 30
                )  # Вычисляем количество пробелов, чтобы выровнять текст.

                text += key  # Добавляем ключ к тексту.

                if pad > 0:  # Если нужно добавить пробелы для выравнивания,
                    text += " " * pad  # добавляем их.

                if total_len > 30:  # Если общая длина превышает 30 символов,
                    text += " " * 2  # добавляем 2 дополнительных пробела.

                text += str(v)  # Добавляем значение к тексту.
                text += "\n"  # Добавляем перевод строки между ключами и значениями.
            text += "\n"  # Добавляем пустую строку после каждого объекта.
            text += "******************************"  # Добавляем разделительную строку.
            text += "\n"

        text += ""  # Пустая строка (это выглядит как ошибка, потому что она ничего не делает).
        index = 0  # Начальный индекс для разделения текста на части.
        size = 4000  # Максимальная длина каждой части сообщения.
        while len(text) > 0:  # Пока есть текст для обработки:
            part = text[
                index : index + size
            ]  # Выбираем часть текста длиной не более 4000 символов.
            index = part.rfind(
                "\n"
            )  # Находим последний символ перевода строки в части.
            if index == -1:  # Если символ перевода строки не найден,
                index = len(text)  # используем всю часть текста.
            part = text[
                0:index
            ]  # Выбираем часть текста до найденного символа перевода строки.
            messages.append(
                "```\n" + part + "\n```"
            )  # Добавляем часть текста в список сообщений,
            text = text[index:].strip()  # и удаляем ее из исходного текста.

        logger.debug("Finished formatting messages")
    else:
        logger.debug("Input object is empty")

    return messages  # Возвращаем список отформатированных сообщений.


def format_message_list5(obj):
    text = ""  # Создаем пустую строку для хранения текста сообщений.
    messages = []  # Создаем пустой список для хранения отформатированных сообщений.

    if len(obj) > 0:  # Проверяем, есть ли объекты в списке.
        logger.debug("Processing non-empty object")
        for i in obj:  # Проходим по каждому объекту в списке.
            for k, v in i.items():  # Проходим по каждой паре ключ-значение в объекте.
                key = str(k)  # Преобразуем ключ в строку.
                val = str(v)  # Преобразуем значение в строку.
                total_len = len(key) + len(
                    val
                )  # Вычисляем общую длину ключа и значения.
                pad = (
                    30 - total_len % 30
                )  # Вычисляем количество пробелов, чтобы выровнять текст.

                text += key  # Добавляем ключ к тексту.

                if pad > 0:  # Если нужно добавить пробелы для выравнивания,
                    text += " " * pad  # добавляем их.

                if total_len > 30:  # Если общая длина превышает 30 символов,
                    text += " " * 2  # добавляем 2 дополнительных пробела.

                text += str(v)  # Добавляем значение к тексту.
                text += "\n"  # Добавляем перевод строки между ключами и значениями.
            text += "\n"  # Добавляем пустую строку после каждого объекта.
            text += "******************************"  # Добавляем разделительную строку.
            text += "\n"

            text += ""
            index = 0  # Начальный индекс для разделения текста на части.
            size = 4000  # Максимальная длина каждой части сообщения.
            # while len(text) > 0:  # Пока есть текст для обработки:
            part = text[
                index : index + size
            ]  # Выбираем часть текста длиной не более 4000 символов.
            index = part.rfind(
                "\n"
            )  # Находим последний символ перевода строки в части.
            if index == -1:  # Если символ перевода строки не найден,
                index = len(text)  # используем всю часть текста.
            part = text[
                0:index
            ]  # Выбираем часть текста до найденного символа перевода строки.
            messages.append(
                "```\n" + part + "\n```"
            )  # Добавляем часть текста в список сообщений,
            text = text[index:].strip()  # и удаляем ее из исходного текста.
        logger.debug("Finished formatting messages")
    else:
        logger.debug("Input object is empty")

    return messages  # Возвращаем список отформатированных сообщений.


def xls_to_json_format_change(downloaded_file):
    try:
        logger.info("Начало преобразования XLS в JSON")

        # Создаем буферизированный объект для чтения из загруженного файла
        file_buffer = io.BytesIO(downloaded_file)

        # Открываем книгу Excel
        book = load_workbook(file_buffer)

        # Получаем активный лист из книги Excel
        ws = book.active

        my_list = []  # Создаем пустой список для хранения словарей

        # Находим номер последнего столбца и строки
        last_column = len(list(ws.columns))
        last_row = len(list(ws.rows))

        # Проходимся по каждой строке в таблице Excel
        for row in range(1, last_row + 1):
            my_dict = {}  # Создаем пустой словарь для текущей строки
            # Проходимся по каждому столбцу в текущей строке
            for column in range(1, last_column + 1):
                column_letter = get_column_letter(
                    column
                )  # Получаем буквенное обозначение столбца

                if row > 1:  # Пропускаем первую строку, так как это заголовки
                    # Добавляем элементы в словарь в формате "значение заголовка: значение ячейки"
                    my_dict[str(ws[column_letter + str(1)].value)] = ws[
                        column_letter + str(row)
                    ].value
            if len(my_dict) > 0:  # Убеждаемся, что словарь не пустой
                my_list.append(my_dict)  # Добавляем словарь в список
        logger.info("Преобразование завершено")
        return my_list  # Возвращаем список словарей
    except openpyxl.utils.exceptions.InvalidFileException:
        logger.error("Загруженный файл не является действительным файлом Excel.")
        return None
    except Exception as e:
        logger.error(f"Произошла ошибка: {e}")
        traceback.print_exc()
        return None


def filter_info(info_):
    # Создаем пустую строку для хранения отфильтрованных элементов
    filtered_str = " ".join(
        [
            # Проходим по каждому элементу списка info_
            item
            # Отбираем элементы, которые содержат хотя бы одну букву или символ "№"
            for item in info_
            if any(char.isalpha() or char == "№" for char in item)
            # Отбираем элементы, которые не содержат строки "БЕЗ" или "НДС"
            and item not in ["БЕЗ", "НДС"]
        ]
    )
    logger.info("Фильтрация завершена")
    return filtered_str


def find_INN(resource, query, API_KEY):
    logger.info("Поиск ИНН по запросу")
    # Функция для поиска ИНН по запросу с использованием API Dadata
    BASE_URL = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/suggest/"
    url = BASE_URL + resource
    headers = {
        "Authorization": "Token " + API_KEY,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    data = {"query": query}
    # Отправляем POST-запрос к API для поиска ИНН
    res = requests.post(url, data=json.dumps(data), headers=headers)
    # Возвращаем JSON-ответ
    logger.info("Поиск ИНН завершен")
    return res.json()


def process_PDF_files(downloaded_zip_file):
    logger.info("Начало обработки PDF файлов")

    API_KEY = "e847ba51dfe5006957aca33cbd3e158f234b2bfe"
    try:
        logger.debug("Начало обработки zip-файла с PDF файлами")
        with zipfile.ZipFile(io.BytesIO(downloaded_zip_file), "r") as zip_ref:
            results = []  # Здесь будем хранить результаты обработки
            dict_ = {}
            for file_name in zip_ref.namelist():
                if file_name.lower().endswith(".pdf"):
                    pdf_content = zip_ref.read(file_name)
                    pdf_file = io.BytesIO(pdf_content)
                    try:
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        for page_num in range(len(pdf_reader.pages)):
                            page = pdf_reader.pages[page_num]
                            page_content = page.extract_text()
                            page_content = page_content.split("\n")
                            name_ = ""
                            phone_ = ""
                            address_ = ""
                            inn_ = ""
                            сompany_ = ""
                            product_ = ""
                            address__ = (
                                False  # Флаг для обозначения начала сбора адреса
                            )

                            for line in page_content:
                                if re.search(r"^ИНН:", line):
                                    inn_ = line.replace("ИНН: ", "").replace(".", "")
                                    if inn_[:-1] in dict_.keys():
                                        сompany_ = dict_[inn_[:-1]]
                                    else:
                                        name_company = find_INN(
                                            "party", inn_[:-1], API_KEY
                                        )["suggestions"][0]["value"]
                                        dict_[inn_[:-1]] = name_company
                                        сompany_ = dict_[inn_[:-1]]
                                if re.search(r"^ФИО:", line):
                                    name_ = line.replace("ФИО: ", "").replace(".", "")
                                if re.search(r"^Телефон:", line):
                                    if phone_ == "":
                                        phone_ = line.replace("Телефон: ", "")
                                        if phone_[0] == "7":
                                            phone_ = "8" + phone_[1:]
                                if address__:
                                    address_ = (
                                        address_
                                        + " "
                                        + line.replace("Адрес: ", "")
                                        .replace(" ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                        .replace("ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                    )
                                    if "ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ" in line:
                                        address__ = False
                                if re.search(r"^Адрес:", line) or re.search(
                                    r"^ПВЗ:", line
                                ):
                                    if not "ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ" in line:
                                        address__ = True
                                    address_ = (
                                        line.replace("Адрес: ", "")
                                        .replace(" ПВЗ: ", "")
                                        .replace("ПВЗ: ", "")
                                        .replace(" ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                        .replace("ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                    )
                                if "Арт." in line:
                                    info_ = line.split(" ")
                                    if len(info_) > 2:
                                        if len(info_[2]) > 1:
                                            product_ = filter_info(info_)

                                        else:
                                            info_2 = (
                                                line.replace(")", "__")
                                                .replace("(", "__")
                                                .split("__")
                                            )
                                            product_ = filter_info(info_2)

                            result_entry = {
                                "ФИО": name_.strip(),
                                "Телефон": phone_.strip(),
                                "Адрес": address_.strip(),
                                "Продукт": product_.strip(),
                                "ИНН": inn_.strip(),
                                "Компания": сompany_.strip(),
                            }

                            results.append(result_entry)
                    except Exception as e:
                        logger.error(f"Ошибка при чтении PDF-файла: {file_name}")
                        traceback.print_exc()

        # Преобразование результатов в формат JSON
        # json_results = json.dumps(results, ensure_ascii=False, indent=4)
        logger.info("Обработка PDF файлов завершена")

        return results

    except zipfile.BadZipFile:
        logger.error("Ошибка: Неверный zip-файл.")
        traceback.print_exc()
    except Exception as e:
        logger.error(f"Ошибка: {str(e)}")
        traceback.print_exc()


def process_PDF_files_rar(downloaded_rar_file):
    logger.info("Начало обработки RAR файлов")
    API_KEY = "e847ba51dfe5006957aca33cbd3e158f234b2bfe"
    try:
        with rarfile.RarFile(io.BytesIO(downloaded_rar_file), "r") as rar_ref:
            results = []  # Здесь будем хранить результаты обработки
            dict_ = {}
            logger.debug("Начало обработки rar-файла с PDF файлами")
            for file_name in rar_ref.namelist():
                if file_name.lower().endswith(".pdf"):
                    pdf_content = rar_ref.read(file_name)
                    pdf_file = io.BytesIO(pdf_content)
                    try:
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        for page_num in range(len(pdf_reader.pages)):
                            page = pdf_reader.pages[page_num]
                            page_content = page.extract_text()
                            page_content = page_content.split("\n")
                            name_ = ""
                            phone_ = ""
                            address_ = ""
                            inn_ = ""
                            сompany_ = ""
                            product_ = ""
                            address__ = (
                                False  # Флаг для обозначения начала сбора адреса
                            )

                            for line in page_content:
                                if re.search(r"^ИНН:", line):
                                    inn_ = line.replace("ИНН: ", "").replace(".", "")
                                    if inn_[:-1] in dict_.keys():
                                        сompany_ = dict_[inn_[:-1]]
                                    else:
                                        name_company = find_INN(
                                            "party", inn_[:-1], API_KEY
                                        )["suggestions"][0]["value"]
                                        dict_[inn_[:-1]] = name_company
                                        сompany_ = dict_[inn_[:-1]]
                                if re.search(r"^ФИО:", line):
                                    name_ = line.replace("ФИО: ", "").replace(".", "")
                                if re.search(r"^Телефон:", line):
                                    if phone_ == "":
                                        phone_ = line.replace("Телефон: ", "")
                                        if phone_[0] == "7":
                                            phone_ = "8" + phone_[1:]
                                if address__:
                                    address_ = (
                                        address_
                                        + " "
                                        + line.replace("Адрес: ", "")
                                        .replace(" ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                        .replace("ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                    )
                                    if "ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ" in line:
                                        address__ = False
                                if re.search(r"^Адрес:", line) or re.search(
                                    r"^ПВЗ:", line
                                ):
                                    if not "ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ" in line:
                                        address__ = True
                                    address_ = (
                                        line.replace("Адрес: ", "")
                                        .replace(" ПВЗ: ", "")
                                        .replace("ПВЗ: ", "")
                                        .replace(" ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                        .replace("ИНФОРМАЦИЯ ОБ ОТПРАВЛЕНИИ", "")
                                    )
                                if "Арт." in line:
                                    info_ = line.split(" ")
                                    if len(info_) > 2:
                                        if len(info_[2]) > 1:

                                            product_ = filter_info(info_)

                                        else:
                                            info_2 = (
                                                line.replace(")", "__")
                                                .replace("(", "__")
                                                .split("__")
                                            )

                                            product_ = filter_info(info_2)

                            result_entry = {
                                "ФИО": name_.strip(),
                                "Телефон": phone_.strip(),
                                "Адрес": address_.strip(),
                                "Продукт": product_.strip(),
                                "ИНН": inn_.strip(),
                                "Компания": сompany_.strip(),
                            }

                            results.append(result_entry)
                    except Exception as e:
                        logger.error(f"Ошибка при чтении PDF-файла: {file_name}")
                        traceback.print_exc()

        # Преобразование результатов в формат JSON
        # json_results = json.dumps(results, ensure_ascii=False, indent=4)
        logger.info("Обработка RAR файлов завершена")
        return results

    except rarfile.BadRarFile:
        logger.error("Ошибка: Неверный rar-файл.")
        traceback.print_exc()
    except Exception as e:
        logger.error(f"Ошибка: {str(e)}")
        traceback.print_exc()


def process_PDF_files_(downloaded_zip_file):
    logger.info("Начало обработки ZIP файлов")
    try:
        with zipfile.ZipFile(io.BytesIO(downloaded_zip_file), "r") as zip_ref:
            for file_name in zip_ref.namelist():
                if file_name.lower().endswith(".pdf"):
                    pdf_content = zip_ref.read(file_name)
                    pdf_file = io.BytesIO(pdf_content)
                    try:
                        pdf_reader = PyPDF2.PdfReader(pdf_file)
                        text = ""
                        for page_num in range(len(pdf_reader.pages)):
                            page = pdf_reader.pages[page_num]
                            text += page.extract_text()
                    except Exception as e:
                        logger.error(f"Ошибка при чтении PDF-файла: {file_name}")
                        traceback.print_exc()
        logger.info("Обработка ZIP файлов завершена")
        return text
    except zipfile.BadZipFile:
        logger.error("Ошибка: Неверный zip-файл.")
        traceback.print_exc()
    except Exception as e:
        logger.error(f"Ошибка: {str(e)}")
        traceback.print_exc()
def contains_br_code(row):
    """
    Проверяет, содержит ли первый элемент списка подстроку 'БР-'.

    :param row: Список, представляющий строку таблицы.
    :return: True, если первый элемент содержит 'БР-', иначе False.
    """
    return row and isinstance(row[0], str) and "БР-" in row[0]


def process_pdf_files(downloaded_zip_file):
    """
    Обрабатывает PDF-файлы из zip-архива, извлекая строки с подстрокой 'БР-'.

    :param downloaded_zip_file: Содержимое zip-архива в виде байтов.
    :return: Список извлечённых данных.
    """
    logger.info("Начало обработки PDF файлов")

    try:
        logger.debug("Начало обработки zip-файла с PDF файлами")
        with zipfile.ZipFile(io.BytesIO(downloaded_zip_file), "r") as zip_ref:
            results = []  # Здесь будем хранить результаты обработки

            for file_name in zip_ref.namelist():
                if file_name.lower().endswith(".pdf"):
                    logger.debug(f"Обработка файла: {file_name}")
                    pdf_content = zip_ref.read(file_name)
                    pdf_file = io.BytesIO(pdf_content)

                    try:
                        with pdfplumber.open(pdf_file) as pdf:
                            for page_number, page in enumerate(pdf.pages, start=1):
                                logger.debug(
                                    f"Обработка страницы {page_number} файла {file_name}"
                                )
                                tables = page.extract_tables()

                                for table in tables:
                                    for row in table:
                                        if contains_br_code(row):
                                            results.append(
                                                row
                                            )  # Добавляем строку в результаты

                    except Exception as e:
                        logger.error(f"Ошибка при обработке PDF-файла: {file_name}")
                        traceback.print_exc()

        logger.info("Обработка PDF файлов завершена")
        return results

    except zipfile.BadZipFile:
        logger.error("Ошибка: Неверный zip-файл.")
        traceback.print_exc()
    except Exception as e:
        logger.error(f"Ошибка: {str(e)}")
        traceback.print_exc()


def process_pdf_file_no_zip(pdf_content):
    """
    Обрабатывает PDF-файл, извлекая строки с подстрокой 'БР-'.

    :param pdf_content: Содержимое PDF-файла в виде байтов.
    :return: Список извлечённых данных.
    """
    logger.info("Начало обработки PDF файла")

    try:
        results = []  # Список для сохранения извлечённых строк
        pdf_file = io.BytesIO(pdf_content)

        try:
            with pdfplumber.open(pdf_file) as pdf:
                for page_number, page in enumerate(pdf.pages, start=1):
                    logger.debug(f"Обработка страницы {page_number}")
                    tables = page.extract_tables()

                    for table in tables:
                        for row in table:
                            if contains_br_code(row):
                                results.append(row)  # Добавляем строку в результаты

        except Exception as e:
            logger.error("Ошибка при обработке PDF-файла")
            traceback.print_exc()

        logger.info("Обработка PDF файла завершена")
        return results

    except Exception as e:
        logger.error(f"Ошибка: {str(e)}")
        traceback.print_exc()