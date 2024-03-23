from bd.model import Clients
from arrow import utcnow, get
from bd.model import Session, Clients
from .util import xls_to_json_format_change
from pprint import pprint
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, get_column_letter


name = "Загрузить данные"
desc = "Загружает данне из xls в базу"
mime = "text"


class FileInput:
    name = "Файл"
    desc = "Отправте файл в формате xls"
    type = "FILE"


def get_inputs(session: Session):
    return {"file": FileInput}


def generate(session: Session):
    params = session.params["inputs"]["0"]
    data = "/Users/dmitrijsuvalov/Timur/{}".format(params["file"])
    wb = load_workbook(filename=data)

    result = xls_to_json_format_change(wb)

    for item in result:
        item["closeDate"] = utcnow().shift(hour=3).isoformat()
        Clients.objects(Телефон=item["Телефон"]).update(**item, upsert=True)

    return result
