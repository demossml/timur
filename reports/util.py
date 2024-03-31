from bd.model import Session
from arrow import utcnow, get
from typing import List, Tuple
from pprint import pprint


from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Принимает словарь с данными о продукте


def period_to_date(period: str) -> utcnow:
    """
    :param period: day, week,  fortnight, month, two months,
    :return: utcnow - period
    """
    if period == "day":
        return utcnow().to("local").replace(hour=3, minute=00).isoformat()
    if period == "week":
        return (
            utcnow().to("local").shift(days=-7).replace(hour=3, minute=00).isoformat()
        )
    if period == "fortnight":
        return (
            utcnow().to("local").shift(days=-14).replace(hour=3, minute=00).isoformat()
        )
    if period == "month":
        return (
            utcnow().to("local").shift(months=-1).replace(hour=3, minute=00).isoformat()
        )
    if period == "two months":
        return (
            utcnow().to("local").shift(months=-2).replace(hour=3, minute=00).isoformat()
        )
    if period == "6 months":
        return (
            utcnow().to("local").shift(months=-6).replace(hour=3, minute=00).isoformat()
        )
    if period == "12 months":
        return (
            utcnow()
            .to("local")
            .shift(months=-12)
            .replace(hour=3, minute=00)
            .isoformat()
        )
    if period == "24 months":
        return (
            utcnow()
            .to("local")
            .shift(months=-24)
            .replace(hour=3, minute=00)
            .isoformat()
        )
    if period == "48 months":
        return (
            utcnow()
            .to("local")
            .shift(months=-48)
            .replace(hour=3, minute=00)
            .isoformat()
        )
    raise Exception("Period is not supported")


def period_to_date_2(period: str) -> utcnow:
    """
    :param period: day, week,  fortnight, month, two months,
    :return: utcnow + period
    """
    if period == "day":
        return utcnow().replace(hour=3, minute=00).isoformat()
    if period == "week":
        return utcnow().shift(days=7).replace(hour=3, minute=00).isoformat()
    if period == "fortnight":
        return utcnow().shift(days=14).replace(hour=3, minute=00).isoformat()
    if period == "month":
        return utcnow().shift(months=1).replace(hour=3, minute=00).isoformat()
    if period == "two months":
        return utcnow().shift(months=2).replace(hour=3, minute=00).isoformat()
    if period == "6 months":
        return utcnow().shift(months=6).replace(hour=3, minute=00).isoformat()
    if period == "12 months":
        return utcnow().shift(months=12).replace(hour=3, minute=00).isoformat()
    if period == "24 months":
        return utcnow().shift(months=24).replace(hour=3, minute=00).isoformat()
    if period == "48 months":
        return utcnow().shift(months=48).replace(hour=3, minute=00).isoformat()
    raise Exception("Period is not supported")


def get_intervals(
    min_date: str, max_date: str, unit: str, measure: float
) -> List[Tuple[str, str]]:
    """
    :param min_date: дата начала пириода
    :param max_date: дата окончания пириода
    :param unit: days, weeks,  fortnights, months
    :param measure: int шаг
    :return: List[Tuple[min_date, max_date]]
    """
    output = []
    while min_date < max_date:
        # записывет в перменную temp минимальную дату плюс (unit: measure)
        temp = get(min_date).shift(**{unit: measure}).isoformat()
        # записывает в output пару дат min_date и  меньшую дату min_date max_date или temp
        output.append((min_date, min(temp, max_date)))
        # меняет значение min_date на temp
        min_date = temp
    return output


def get_period(session: Session):
    """
    :param session:
    :return: {'since': str, 'until': str}
    """
    period_in = ["day", "week", "fortnight", "month"]
    if session.params["inputs"]["0"]["period"] not in period_in:
        return {
            "since": get(session.params["inputs"]["0"]["openDate"])
            .replace(day=1)
            .isoformat(),
            "until": get(session.params["inputs"]["0"]["openDate"])
            .ceil("month")
            .isoformat(),
        }
    if session.params["inputs"]["0"]["period"] == "day":
        return {
            "since": period_to_date(session.params["inputs"]["0"]["period"]),
            "until": utcnow().isoformat(),
        }

    else:
        return {
            "since": get(session.params["inputs"]["0"]["openDate"])
            .replace(hour=3, minute=00)
            .isoformat(),
            "until": get(session.params["inputs"]["0"]["closeDate"])
            .replace(hour=23, minute=00)
            .isoformat(),
        }


def get_period_day(session: Session):
    """
    :param session:
    :return: {'since': str, 'until': str}
    """
    if session.params["inputs"]["0"]["period"] == "day":
        return {
            "since": period_to_date(session.params["inputs"]["0"]["period"]),
            "until": get(period_to_date(session.params["inputs"]["0"]["period"]))
            .replace(hour=23, minute=00)
            .isoformat(),
        }

    else:
        return {
            "since": get(session.params["inputs"]["0"]["openDate"])
            .replace(hour=0, minute=1)
            .isoformat(),
            "until": get(session.params["inputs"]["0"]["openDate"])
            .replace(hour=23, minute=59)
            .isoformat(),
        }


# Получает список славарей
# # Отдает xls
# def json_to_xls_format_change_(list: list):
#     # Создаем новую книгу Excel
#     book = Workbook()

#     # Выбираем активный лист в книге
#     sheet = book.active

#     # Задаем список, указывающий, в каких ячейках будут располагаться названия столбцов
#     sheet_row = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "J1", "K1", "L1", "M1"]

#     # Создаем список для хранения названий столбцов
#     columns_name = []

#     # Проходим по каждому элементу во входном списке
#     for item in list:
#         # Проходим по ключам и значениям элемента
#         for k, v in item.items():
#             # Если название ключа еще не встречалось, добавляем его в список названий столбцов
#             if k not in columns_name:
#                 columns_name.append(k)

#     # Переменная для отслеживания текущего столбца
#     columns = 0

#     # Записываем названия столбцов в соответствующие ячейки
#     for name in columns_name:
#         sheet[sheet_row[columns]] = name
#         columns += 1

#     # Переменная для отслеживания текущей строки
#     row = 2

#     # Записываем данные из входного списка в ячейки, начиная со второй строки
#     for item in list:
#         # Проверяем, что элемент не пустой
#         if len(item) > 0:
#             last_column = 0
#             # Проходим по ключам и значениям элемента
#             for k, v in item.items():
#                 # Записываем значение в соответствующую ячейк
#                 sheet[row][last_column].value = v
#                 last_column += 1
#         row += 1  # Переходим к следующей строке

#     # Возвращаем созданную книгу Excel
#     return book


def xls_to_json_format_change(book):

    # Получаем активный лист из книги Excel
    ws = book.active

    my_list = []  # Создаем пустой список для хранения словарей

    # Находим номер последнего столбца и строки
    last_column = len(list(ws.columns))
    last_row = len(list(ws.rows))

    # Проходимся по каждой строке в таблице Excel
    for row in range(1, last_row + 1):
        my_dict = {}  # Создаем пустой словарь для текущей строки
        # Проходимся по каждому столбцу в текущей строке
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(
                column
            )  # Получаем буквенное обозначение столбца
            if row > 1:  # Пропускаем первую строку, так как это заголовки
                # Добавляем элементы в словарь в формате "значение заголовка: значение ячейки"
                my_dict[ws[column_letter + str(1)].value] = ws[
                    column_letter + str(row)
                ].value
        if len(my_dict) > 0:  # Убеждаемся, что словарь не пустой
            my_list.append(my_dict)  # Добавляем словарь в список
    return my_list  # Возвращаем список словарей


def he_she(data_list: list):
    """
    Разделяет список данных на два списка: один для мужчин, другой для женщин.

    Args:
        data_list (list): Список словарей с данными. Каждый словарь содержит информацию о человеке, включая ФИО.

    Returns:
        tuple: Кортеж из двух списков: первый список содержит данные о мужчинах, второй — о женщинах.
            Первый элемент кортежа — список мужчин.
            Второй элемент кортежа — список женщин.
            Третий элемент кортежа — количество удаленных дубликатов для мужчин.
            Четвертый элемент кортежа — количество удаленных дубликатов для женщин.
    """
    he = []  # Список для мужчин
    she = []  # Список для женщин

    # Создаем множество для хранения уникальных номеров телефонов
    phones_she = set()
    phones_he = set()

    deleted_duplicates_he = 0
    deleted_duplicates_she = 0

    for item in data_list:
        try:
            if item["ФИО"][-1].upper() == "А":
                if item["Телефон"] not in phones_she:
                    she.append(item)  # Добавляем женщину в список
                else:
                    deleted_duplicates_she += 1
                phones_she.add(item["Телефон"])  # Иначе добавляем телефон в множество
            else:
                if item["Телефон"] not in phones_he:
                    he.append(item)  # Добавляем мужчину в список
                else:
                    deleted_duplicates_he += 1
                phones_he.add(item["Телефон"])  # Иначе добавляем телефон в множество
        except IndexError:
            print("Ошибка: Неправильный формат данных ФИО для элемента:", item)

    return he, she, deleted_duplicates_he, deleted_duplicates_she


def json_to_xls_format_change(data_list: list, gender: str, number_of_lins: int):
    # Создаем новую книгу Excel
    book = Workbook()

    # Выбираем активный лист в книге
    sheet = book.active

    # Используем множество для хранения уникальных названий столбцов
    columns_name = set()

    # Создаем множество для хранения уникальных номеров телефонов
    # phones = set()

    # Проходим по каждому элементу во входном списке
    for item in data_list:
        # Добавляем названия столбцов в множество
        columns_name.update(item.keys())

    # Записываем названия столбцов в первую строку
    for col_idx, column_name in enumerate(columns_name, start=1):
        sheet.cell(row=1, column=col_idx, value=column_name)

    # Записываем данные из входного списка в ячейки и удаляем дубликаты по столбцу "Телефон"
    deleted_duplicates = 0

    # Записываем данные из входного списка в ячейки
    for row_idx, item in enumerate(data_list, start=2):
        # Проходим по названиям столбцов
        for col_idx, column_name in enumerate(columns_name, start=1):
            # Записываем значение в соответствующую ячейку, если оно есть в словаре
            sheet.cell(row=row_idx, column=col_idx, value=item.get(column_name))

        # # Получаем значение из столбца "Телефон" текущей строки
        # phone = item.get("Телефон")
        # if phone in phones:  # Если телефон уже встречался, удаляем строку
        #     sheet.delete_rows(row_idx)
        #     deleted_duplicates += 1
        # else:
        #     phones.add(phone)  # Иначе добавляем телефон в множество

    # Возвращаем созданную книгу Excel и количество удаленных дубликатов
    return book, {
        f"Выгружено строк {gender}": sheet.max_row - 1,
        "Кол. удал. дубл.": number_of_lins,
    }
